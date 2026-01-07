import openpyxl
import json
import os

excel_name = "Tabelas_Sondagem_SMIT_CRAO.xlsx"
output_file = "extracted_tables.json"

# Map Sheet Name -> Tank ID in SoundingData.js
SHEET_MAP = {
    "FO_DAY.P": "FO_DAY_P",
    "FO_DAY.S": "FO_DAY_S",
    "FO_AFT.P": "FO_DB_AFT_P",
    "FO_AFT.S": "FO_DB_AFT_S",
    "FO_DB.P":  "FO_DB_P",
    "FO_DB.S":  "FO_DB_S",
    "FO_DB.C":  "FO_DB_C",
    "FO_OF.P":  "FO_OF_P"
}

def extract_tables():
    if not os.path.exists(excel_name):
        print(f"File {excel_name} not found.")
        return

    try:
        wb = openpyxl.load_workbook(excel_name, data_only=True)
        tables = {}

        for sheet_name, tank_id in SHEET_MAP.items():
            if sheet_name in wb.sheetnames:
                print(f"Processing {sheet_name} -> {tank_id}")
                ws = wb[sheet_name]
                data_points = []
                
                # Assuming data starts at row 6 (based on inspection showing headers at row 5)
                # Columns: A=Sounding(mm), B=Volume(litros) -> index 0 and 1
                for row in ws.iter_rows(min_row=6, values_only=True):
                    if row[0] is not None and row[1] is not None:
                        try:
                            s = float(row[0])
                            v = float(row[1])
                            data_points.append({"s": s, "v": v})
                        except ValueError:
                            continue
                
                # Structure for SoundingData.js: { "0": [...] }
                tables[tank_id] = { "0": data_points }
            else:
                print(f"Warning: Sheet {sheet_name} not found.")

        # Save to JSON file for easy copy-paste
        with open(output_file, "w", encoding="utf-8") as f:
            json.dump(tables, f, indent=4)
        print(f"Extraction complete. Data saved to {output_file}")

    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    extract_tables()
