import openpyxl

excel_name = "Tabelas_Sondagem_SMIT_CARAJA.xlsx"
sheet_name = "FO_DB.C"

def inspect_sheet():
    try:
        wb = openpyxl.load_workbook(excel_name, data_only=True)
        if sheet_name not in wb.sheetnames:
            print(f"Sheet {sheet_name} not found.")
            return

        ws = wb[sheet_name]
        print(f"--- Raw Data Dump for {sheet_name} ---")
        for row in ws.iter_rows(values_only=True):
            # Print all rows to see structure, formatting, etc.
            # Filter out completely empty rows
            if any(cell is not None for cell in row):
                print(row)

    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    inspect_sheet()
