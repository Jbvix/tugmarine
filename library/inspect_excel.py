import openpyxl
import os

excel_name = "Tabelas_Sondagem_SMIT_CARAJA.xlsx"

if os.path.exists(excel_name):
    try:
        wb = openpyxl.load_workbook(excel_name, data_only=True)
        with open("excel_inspection_utf8.txt", "w", encoding="utf-8") as f:
            f.write(f"Sheets: {wb.sheetnames}\n")
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                f.write(f"--- Sheet: {sheet_name} ---\n")
                # Print first 5 rows to identify headers
                for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                    f.write(f"{row}\n")
    except Exception as e:
        print(f"Error: {e}")
else:
    print(f"File {excel_name} not found.")
